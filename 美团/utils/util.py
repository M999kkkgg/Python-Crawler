import fake_useragent
import requests
import random
import time
import jieba
import re
import openpyxl
import math
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib import cm
from utils.utils_dbase import MySqlite
import numpy as np
import string


def pause():
    pause_temp = input("暂停")


def download_img(img_url, path):
    if img_url == '':
        print("无url")
        return
    img_url = img_url.replace("/w.h/", "/")
    # 代理的ip
    # proxy_list = []
    # with open('ip.json', 'r', encoding='utf-8') as f:
    #     ips_dict = json.load(f)['data']
    # for ip in ips_dict:
    #     dic = {'http': str(ip['ip']) + ':' + str(ip['port'])}
    #     proxy_list.append(dic)
    # proxy = random.choice(proxy_list)
    # 随机的浏览器
    header = {}
    ua = fake_useragent.UserAgent()
    header['User-Agent'] = str(ua.random)
    # r = requests.get(img_url, headers=header, proxies=proxy)
    try:
        r = requests.get(img_url, headers=header)
    except :
        pass
    if r.status_code == 200:
        open(path, 'wb').write(r.content)
    else:
        print("保存失败: " + path)
    r.close()
    # 随机休眠1-3秒来模拟真实访问
    seconds = random.randint(1, 3)
    print(f"随机休眠{seconds}s...")
    time.sleep(seconds)


def separate_words(sentence, stop_words):
    """
    分词函数，会去除标点符号进行分析
    采用jieba.lcut函数

    :param sentence: 句子
    :param stop_words: 停用词
    :return: 词语[]
    """
    if sentence == '':
        return ['无']
    punc = 'qwertyuiopasdfghjklzxcvbnmQWERTYUIOPLKJHGFDSAZXCVBNM0123456789~`!#$%^&*()_+-=|\';":/.,?><~·！@#￥%……&*（）——+-=“”：‘’；、。，？》《{}'
    sentence_re = re.sub(r"[%s]+" % punc, "", sentence)
    words = jieba.lcut(sentence_re, cut_all=True)
    # 去除停用词
    new_words = []
    for word in words:
        if word in stop_words:
            continue
        else:
            new_words.append(word)
    return new_words


def get_TF(words):
    words_set = set(words)
    word_num = {}
    for word_set in words_set:
        word_num[word_set] = 0
    i = 0
    num_words = len(words)  # 总词数
    for word_set in words_set:
        word_num[word_set] += 1
    # 计算每个词的TF
    word_TF = {}
    for word_set in words_set:
        word_TF[word_set] = word_num[word_set] / float(num_words)
    return word_TF, words_set



def get_IDF(places_words, words_set):
    # 景点总数
    K = len(places_words)
    # 计算包含某个词的景点数目
    in_num = {}
    word_IDF = {}
    for word_set in words_set:
        in_num[word_set] = 1
        word_IDF[word_set] = 0.0
    for word_set in words_set:
        for place_words in places_words:
            if word_set in place_words[2]:
                in_num[word_set] += 1
    for word_set in words_set:
        temp = math.log(float(K)/in_num[word_set], 10)
        word_IDF[word_set] = temp
    return word_IDF



def get_place_infor_from_city(read_path, city):
    """
    读取存储景点信息的xlsx文件，得到景点部分信息
    一共1197个景点

    :param read_path: xlsx文件存储位置
    :param city: 待读取的城市（这里传入[id, 中文, 英文]）
    :return: [城市名, 景点名, 景点一级分类, 景点简介]
    """
    # 打开工作表
    wb = openpyxl.load_workbook(read_path)
    sheet = wb[city[1]]
    # print(f"正在获取 {city[1]} 的景点信息...")
    p_informs = []
    i = 0
    for names, types, informs in zip(sheet['A'], sheet['B'], sheet['L']):
        # 忽略第一行
        if i == 0:
            i = 1
            continue
        p_name = str(names.value)
        p_type = str(types.value)
        p_inform = str(informs.value)
        p_informs.append([city[2], p_name, p_type, p_inform])
    wb.close()
    return p_informs


def get_food_infor_from_city(read_path, city):
    # 打开工作表
    wb = openpyxl.load_workbook(read_path)
    sheet = wb[city[1]]
    f_informs = []
    i = 0
    for names, types in zip(sheet['A'], sheet['B']):
        # 忽略第一行
        if i == 0:
            i = 1
            continue
        f_name = str(names.value)
        f_type = str(types.value)
        f_informs.append([city[2], f_name, f_type])
    wb.close()
    return f_informs


def read_hotel_xlsx(read_path, city_names_ch):
    wb = openpyxl.load_workbook(read_path)
    hotel_data = []
    # '酒店名称', '酒店地址', '酒店图片链接', '酒店所属区域', '纬度', '经度', '酒店类型', '酒店价格', '销售数', '评分', '标签', '可提供的服务', '城市'
    for city_name_ch in city_names_ch:
        sheet = wb[city_name_ch]
        temp_data = []
        i = 0
        for row in sheet.rows:
            # 跳过第一行
            if i == 0:
                i = 1
                continue
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            row_data.append(city_name_ch)
            temp_data.append(row_data)
        hotel_data.extend(temp_data)
    return hotel_data


def read_food_xlsx(read_path, city_names_ch):
    wb = openpyxl.load_workbook(read_path)
    food_data = []
    # 店名	店类型	图片	所属区域	均价	评分	纬度	经度	标签	卖点 城市
    for city_name_ch in city_names_ch:
        sheet = wb[city_name_ch]
        temp_data = []
        i = 0
        for row in sheet.rows:
            # 跳过第一行
            if i == 0:
                i = 1
                continue
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            row_data.append(city_name_ch)
            temp_data.append(row_data)
        food_data.extend(temp_data)
    return food_data


def read_place_xlsx(read_path, city_names_ch):
    wb = openpyxl.load_workbook(read_path)
    place_data = []
    # 景点名	景点类型	景点地址	所属区域	URL	平均分	纬度	经度	销售量	最低价	平均价	景区简介 城市
    for city_name_ch in city_names_ch:
        sheet = wb[city_name_ch]
        temp_data = []
        i = 0
        for row in sheet.rows:
            # 跳过第一行
            if i == 0:
                i = 1
                continue
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            row_data.append(city_name_ch)
            temp_data.append(row_data)
        place_data.extend(temp_data)
    return place_data



class MplColorHelper:

    def __init__(self, cmap_name, start_val, stop_val):
        self.cmap_name = cmap_name
        self.cmap = plt.get_cmap(cmap_name)
        self.norm = mpl.colors.Normalize(vmin=start_val, vmax=stop_val)
        self.scalarMap = cm.ScalarMappable(norm=self.norm, cmap=self.cmap)

    def get_rgb(self, val):
        return self.scalarMap.to_rgba(val)


def get_place_infor_from_user(sq_path, point_num, place_like, city_want, standard):
    my_sqlite = MySqlite(sq_path)
    my_sqlite.link_to_sqlite()
    places_infor = []
    for item in place_like:
        temp = my_sqlite.get_place_infor_from_cityName_and_typeName(city_want, item, standard)
        places_infor.extend(list(temp))
    # 如果符合用户喜欢的景点数目不足，则按照standard随机挑选景点
    while len(places_infor) < point_num:
        temp = None
        while True:
            temp = my_sqlite.get_place_infor_from_random(city_want, standard)
            temp = list(temp)
            if temp not in places_infor:
                break
        places_infor.extend(temp)

    # 如果符合用户喜欢的景点数目过多，则选取前面的（结果已经根据standard排好了序）
    if len(places_infor) > point_num:
        temp_infor = []
        selected_index = []
        while True:
            if len(temp_infor) == point_num:
                break
            index = random.randint(0, len(places_infor) - 1)
            if index in selected_index:
                continue
            temp_infor.append(places_infor[index])
            selected_index.append(index)
        return temp_infor
    else:
        return places_infor


def sphere_distance(p1, p2):
    """
        计算两点p1, p2之间的距离
        p1：（纬度、经度）
        p2: （纬度、经度）
    """
    if p1[0] < 0 or p2[0] < 0 or p1[0] > 90 or p2[0] > 90 or p1[1] < 0 or p2[1] < 0 or p1[1] > 180 or p2[1] > 180:
        return 'Parameter Error.'
    r = 6371
    i1 = (p1[0] / 180) * math.pi
    j1 = (p1[1] / 180) * math.pi
    i2 = (p2[0] / 180) * math.pi
    j2 = (p2[1] / 180) * math.pi
    d = 2 * r * math.asin(
        math.sqrt((math.sin((i2 - i1) / 2)) ** 2 + math.cos(i1) * math.cos(i2) * (math.sin((j2 - j1) / 2)) ** 2))
    return d


def get_distance_from_twoCitys(city1, city2, sq_path):
    my_sq = MySqlite(sq_path)
    my_sq.link_to_sqlite()
    location1 = my_sq.get_location_from_cityName(city1)
    location2 = my_sq.get_location_from_cityName(city2)
    distance = sphere_distance(location1, location2)
    return distance

